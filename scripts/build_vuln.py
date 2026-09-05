# -*- coding: utf-8 -*-
"""취약 3종(외국인주민·등록장애인·기초수급) 시군구 원천 파일 → data/ref/sgg_vuln.json
원천(data/src/vuln/, 연 1회 수동 갱신):
  - 행안부 「2024 지방자치단체 외국인주민 현황」 통계표 xlsx ('1-2. 유형 및 지역별(시.군.구)') — 2024-11-01 기준
  - 복지부 「장애인 등록 현황(2024년말 기준)」 xlsx ('2-3.시군구 연령별 장애인(장애정도)현황') — 2024-12-31 기준
  - (기초수급) 한국사회보장정보원 복지사업 시군구별 수급권자 현황 — 확보 시 basic 필드 채움
비율 분모 = 행안부 주민등록 2026-07(sgg_demo.json pop). 파일의 총인구(2024-11)와 시점이 다르므로 비율은 근사이며 메타에 명시.
매칭 = (시도, 시군구명). 시도명은 2024 체계 → 앱 체계(전남·광주 통합 12) 매핑. 실패 명단은 표준출력으로 보고(지어내지 않음)."""
import json, os, io, re, collections, sys
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
P = lambda *a: os.path.join(ROOT, *a)
V = P("data", "src", "vuln")
import openpyxl, warnings; warnings.filterwarnings("ignore")
idx = json.load(open(P("data", "admin", "sgg_index.json"), encoding="utf-8"))
demo = json.load(open(P("data", "ref", "sgg_demo.json"), encoding="utf-8"))
SIDO_MAP = {"광주광역시": "12", "전라남도": "12", "전남광주통합특별시": "12"}
sido_by_name = {s["sido_name"]: s["sido"] for s in idx}
def sido_code(name):
    n = (name or "").strip()
    if n in SIDO_MAP: return SIDO_MAP[n]
    if n in sido_by_name: return sido_by_name[n]
    # 약칭·개칭 보정: '전북특별자치도'↔'전라북도', '강원특별자치도'↔'강원도'
    alt = {"전라북도": "전북특별자치도", "강원도": "강원특별자치도", "제주도": "제주특별자치도", "세종시": "세종특별자치시"}
    return sido_by_name.get(alt.get(n, ""), None)
by_key = {}
for s in idx: by_key.setdefault((s["sido"], s["name"]), []).append(s["code"])
def find(sido, name):
    name = re.sub(r"\s+", "", name or "")
    c = by_key.get((sido, name))
    if c and len(c) == 1: return c[0]
    # 통합시 소속 구가 '수원시장안구'처럼 붙어 오는 경우 → 마지막 '구' 이름으로
    m = re.match(r"^(.+?시)(.+구)$", name)
    if m: c = by_key.get((sido, m.group(2)));
    return c[0] if c and len(c) == 1 else None
out = {}; unmatched = collections.defaultdict(list); sidos_seen = set()
def num(v):
    try: return float(str(v).replace(",", ""))
    except Exception: return None
# ── 외국인주민 ──
fx = [f for f in os.listdir(V) if "외국인주민" in f and f.endswith(".xlsx")]
if fx:
    wb = openpyxl.load_workbook(os.path.join(V, fx[0]), read_only=True)
    ws = [w for w in wb.worksheets if "시.군.구" in w.title][0]
    cur = None
    for row in ws.iter_rows(min_row=8, values_only=True):
        name = row[0]
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if name == "전국": continue
        sc = sido_code(name)
        if sc and (name.endswith(("시", "도")) and (name in SIDO_MAP or name in sido_by_name or name in ("전라북도", "강원도"))):
            cur = sc; sidos_seen.add(name); continue
        if cur is None: continue
        code = find(cur, name) or (by_key.get((cur, "세종특별자치시")) or [None])[0] if name in ("세종시", "세종특별자치시") else find(cur, name)
        if not code: unmatched["foreign"].append(f"{name}({cur})"); continue
        d = out.setdefault(code, {})
        d.update(pop2411=num(row[1]), foreign=num(row[3]), foreign_noncit=num(row[6]), worker=num(row[9]), marriage=num(row[12]))
    print("foreign rows matched", sum(1 for d in out.values() if "foreign" in d), "sidos", len(sidos_seen))
# ── 등록장애인 ──
dx = [f for f in os.listdir(V) if "장애인" in f and f.endswith(".xlsx")]
if dx:
    wb = openpyxl.load_workbook(os.path.join(V, dx[0]), read_only=True)
    ws = [w for w in wb.worksheets if "시군구" in w.title and "연령별" in w.title][0]
    for row in ws.iter_rows(min_row=8, values_only=True):
        sido, sgg, age = (str(row[0] or "").strip(), str(row[1] or "").strip(), str(row[2] or "").strip())
        if age != "소계" or sgg in ("소계", "합계", ""): continue
        sc = sido_code(sido)
        if not sc: unmatched["disabled"].append(f"{sido} {sgg}"); continue
        code = find(sc, sgg) or ((by_key.get((sc, "세종특별자치시")) or [None])[0] if sc == "36" else None)
        if not code: unmatched["disabled"].append(f"{sgg}({sc})"); continue
        d = out.setdefault(code, {}); d["disabled"] = num(row[11]); d["disabled_severe"] = num(row[5])
    print("disabled rows matched", sum(1 for d in out.values() if "disabled" in d))
# ── 기초수급(선택) ──
# 한국사회보장정보원 「복지사업 시군구별 수급권자 현황」 csv(cp949): 사업명·기준년월·시도·시군구·수급권자수·수급가구수
# basic = '기초생활보장(맞춤형급여)' 수급권자수(생계·의료·주거·교육 통합 자격자), basic_hh = 수급가구수
import csv
bx = [f for f in os.listdir(V) if ("수급권자" in f) and f.endswith(".csv")]
if bx:
    raw = open(os.path.join(V, bx[0]), "rb").read()
    txt = None
    for enc in ("utf-8-sig", "cp949"):
        try: txt = raw.decode(enc); break
        except Exception: pass
    basic_asof = None; nb = 0
    for r in csv.DictReader(io.StringIO(txt)):
        if r.get("사업명") != "기초생활보장(맞춤형급여)": continue
        basic_asof = r.get("기준년월")
        sc = sido_code(r.get("시도"))
        if not sc: unmatched["basic"].append(f"{r.get('시도')} {r.get('시군구')}"); continue
        code = find(sc, r.get("시군구")) or ((by_key.get((sc, "세종특별자치시")) or [None])[0] if sc == "36" else None)
        if not code: unmatched["basic"].append(f"{r.get('시군구')}({sc})"); continue
        d = out.setdefault(code, {}); d["basic"] = num(r.get("수급권자수")); d["basic_hh"] = num(r.get("수급가구수")); nb += 1
    print("basic rows matched", nb, "asof", basic_asof)
# ── 비율·분위 ──
for code, d in out.items():
    pop = (demo.get("sgg", demo).get(code) or {}).get("pop") if isinstance(demo, dict) else None
    if not pop:
        pop = d.get("pop2411")
    d["pop_base"] = pop
    for k in ("foreign", "disabled", "basic"):
        if d.get(k) is not None and pop: d[k + "_r"] = round(d[k] / pop, 4)
def pct(vals, q):
    v = sorted(x for x in vals if x is not None); return v[min(len(v) - 1, int(round(q / 100 * (len(v) - 1))))] if v else None
th = {k: {"p75": pct([d.get(k) for d in out.values()], 75), "p80": pct([d.get(k) for d in out.values()], 80), "median": pct([d.get(k) for d in out.values()], 50)} for k in ("foreign_r", "disabled_r", "basic_r")}
meta = {"built": "2026-09-05", "unit": "sgg", "n": len(out), "asof": {"foreign": "2024-11-01 (행안부 외국인주민 현황 2024)", "disabled": "2024-12-31 (복지부 등록장애인 현황)", "basic": "2025-12 (한국사회보장정보원 복지사업 시군구별 수급권자 현황, 기초생활보장 맞춤형급여 수급권자)"},
        "denominator": "행안부 주민등록 2026-07 (sgg_demo.json pop) — 원천 시점과 다르므로 비율은 근사", "thresholds": th,
        "unmatched": {k: v[:40] for k, v in unmatched.items()}}
json.dump({"meta": meta, "sgg": out}, open(P("data", "ref", "sgg_vuln.json"), "w", encoding="utf-8"), ensure_ascii=False, separators=(",", ":"))
print("thresholds", json.dumps(th, ensure_ascii=False))
print("unmatched", {k: (len(v), v[:12]) for k, v in unmatched.items()})
print("demo keys sample", list(demo.keys())[:5] if isinstance(demo, dict) else type(demo))
